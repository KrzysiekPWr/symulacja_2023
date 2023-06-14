import subprocess
from openpyxl import load_workbook, Workbook
import datetime
import time

wb = load_workbook(filename='results_simulation.xlsx')
actual_time = datetime.datetime.now().strftime("%Y_%m_%d__%H_%M_%S")

# This script calls gradle with the arguments passed to it.
# gradle run --args="arg1 arg2 ... argN"

if __name__ == "__main__":
    planetation = 0.032 #skok 0.001
    map_size = 100
    stars_quantity = 2
    black_holes_quantity = 2
    aggresive_civilisations_quantity = 5
    pacifistic_civilisations_quantity = 5
    number_of_eras = 1000
    number_of_repeats = 1

    # check if there is enough planets for all civilisations!

    call_text = 'gradlew run --args="{} {} {} {} {} {} {} {}"'.format(planetation, map_size, stars_quantity, black_holes_quantity,
                                                            aggresive_civilisations_quantity, pacifistic_civilisations_quantity,
                                                            number_of_eras, number_of_repeats)
    

    return_code = subprocess.check_output(call_text, shell=True)

    return_code = return_code.decode("utf-8")
    # print only line begining with "Return code:"
    split_code = return_code.splitlines()
    

    # code for printing results in manner like:
    # typeOfCiv allResourcesOfCiv 
    # ---------------------------------------------
    #result_for_file = ""
    # for i in range(5, 6 + pacifistic_civilisations_quantity + aggresive_civilisations_quantity):
    #     result_for_file += split_code[i] + "\n"
    # result_for_file += "\n"


    # code for printing results in manner like:
    # idOfCiv TypeOfCiv ResourcesGatheredTillThisEra
    # ---------------------------------------------
    result_for_file = ""
    start = False
    for line in split_code:
        if(line == "-" and start == False):
            start = True
            continue
        elif(line == "-" and start == True):
            break

        if(start == True):
            result_for_file += line + "\n"
        
    # create new sheet in excel spreadsheet
    wb.create_sheet(str(actual_time))
    ws = wb[actual_time]
    # append to excel spreadsheet
    for line in result_for_file.splitlines():
        ws.append(line.split(" "))
    wb.save("results_simulation.xlsx")

    print(result_for_file.splitlines()[1:])
    
    with open("results_simulation_course.txt", "a") as myfile:
        myfile.write(result_for_file)
        myfile.write("-------------------\n")
        myfile.close()
    
wb.close();